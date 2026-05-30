import datetime
import json
import re
import sys
from pathlib import Path

import openpyxl


def clean(value):
    if value in (None, "", "-"):
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    return 0


def iso(value):
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    return str(value) if value else ""


def build_history(ws):
    rows = []
    for row in range(20, ws.max_row + 1):
        if not ws.cell(row, 2).value:
            continue
        rows.append(
            {
                "date": iso(ws.cell(row, 1).value),
                "month": iso(ws.cell(row, 2).value)[:7],
                "salary": clean(ws.cell(row, 3).value),
                "passiveIncome": clean(ws.cell(row, 4).value),
                "income": clean(ws.cell(row, 5).value),
                "netWorth": clean(ws.cell(row, 6).value),
                "mom": clean(ws.cell(row, 7).value),
                "yoy": clean(ws.cell(row, 8).value),
                "cash": clean(ws.cell(row, 17).value),
                "funds": clean(ws.cell(row, 22).value),
                "securities": clean(ws.cell(row, 30).value),
                "crypto": clean(ws.cell(row, 35).value),
                "sideCapital": clean(ws.cell(row, 39).value),
                "realEstate": clean(ws.cell(row, 40).value),
                "privateEquity": clean(ws.cell(row, 44).value),
                "liabilities": clean(ws.cell(row, 48).value),
                "loanDelta": clean(ws.cell(row, 49).value),
                "usdTwd": clean(ws.cell(row, 45).value),
            }
        )
    return rows


def update_account_records(ws, data):
    latest_row = ws.max_row
    latest_month = iso(ws.cell(latest_row, 2).value)[:7]
    updated_accounts = 0
    updated_records = 0
    for account in data.get("accounts", []):
        match = re.search(r"-(\d+)$", account.get("id", ""))
        if not match:
            account.setdefault("records", [])
            continue
        col = int(match.group(1))
        records = []
        for row in range(20, ws.max_row + 1):
            month = iso(ws.cell(row, 2).value)[:7]
            if not month:
                continue
            amount = clean(ws.cell(row, col).value)
            currency = account.get("currency", "TWD")
            usd_twd = clean(ws.cell(row, 45).value) or data.get("meta", {}).get("usdTwd", 31.5)
            value_twd = amount if currency == "TWD" or account.get("type") == "liability" else amount * usd_twd
            records.append(
                {
                    "date": iso(ws.cell(row, 1).value) or f"{month}-01",
                    "month": month,
                    "amount": amount,
                    "valueTwd": value_twd,
                }
            )
        account["records"] = records
        latest = next((item for item in reversed(records) if item.get("month") == latest_month), records[-1] if records else None)
        if latest:
            account["amount"] = latest["amount"]
            account["valueTwd"] = latest["valueTwd"]
        updated_accounts += 1
        updated_records += len(records)
    return updated_accounts, updated_records


def main():
    if len(sys.argv) != 3:
        raise SystemExit("Usage: import-workbook.py <workbook.xlsx> <portfolio-data.json>")
    workbook_path = Path(sys.argv[1])
    data_path = Path(sys.argv[2])
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    data = json.loads(data_path.read_text(encoding="utf-8"))
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet = workbook["資產表"]

    data["history"] = build_history(sheet)
    updated_accounts, updated_records = update_account_records(sheet, data)
    data.setdefault("meta", {})["assetTableImportedAt"] = datetime.datetime.now().isoformat(timespec="seconds")
    data_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    print(
        json.dumps(
            {
                "ok": True,
                "history": len(data.get("history", [])),
                "accounts": updated_accounts,
                "records": updated_records,
                "importedAt": data["meta"]["assetTableImportedAt"],
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
